import io
import json
import os
import zipfile
from copy import copy
from datetime import date
from decimal import Decimal
from typing import List, Optional, TYPE_CHECKING

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

if TYPE_CHECKING:
    from app.models.asignacion import Asignacion
    from app.models.usuario import Usuario
    from app.models.viatico import Viatico

TEMPLATE_PATH = os.path.join(
    os.path.dirname(__file__),
    "templates",
    "FORMATO LEGALIZACION VIATICOS.xlsx",
)

MESES_ES = [
    "", "enero", "febrero", "marzo", "abril", "mayo", "junio",
    "julio", "agosto", "septiembre", "octubre", "noviembre", "diciembre"
]

MESES_ABR = [
    "", "Ene", "Feb", "Mar", "Abr", "May", "Jun",
    "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"
]

def dividir_nombres_apellidos(nombre_completo: str) -> tuple[str, str]:
    if not nombre_completo:
        return ("—", "—")
    partes = [p for p in nombre_completo.strip().split() if p]
    if len(partes) == 0:
        return ("—", "—")
    if len(partes) == 1:
        return (partes[0], "—")
    if len(partes) == 2:
        return (partes[0], partes[1])
    if len(partes) == 3:
        return (" ".join(partes[:2]), partes[2])
    mitad = len(partes) // 2
    return (" ".join(partes[:mitad]), " ".join(partes[mitad:]))

def _formato_fecha_anticipo(f: Optional[date]) -> str:
    if not f:
        return "—"
    if isinstance(f, date):
        return f"{MESES_ES[f.month]} {f.day} {f.year}"
    return str(f)

def _formato_fecha_viaje(f: Optional[date]) -> str:
    if not f:
        return "—"
    if isinstance(f, date):
        return f"{f.day:02d}-{MESES_ABR[f.month]}"
    return str(f)

MAPA_CONCEPTOS = {
    "alimentacion": "Alimentación",
    "transporte": "Transporte",
    "hotel": "Hotel",
    "peajes": "Peajes",
    "parqueadero": "Parqueadero",
    "materiales": "Materiales",
    "alquiler_escalera": "Alquiler de escalera",
    "otros": "Otros",
}

def _formatear_concepto(tipo_gasto: Optional[str]) -> str:
    if not tipo_gasto:
        return "—"
    tg = str(tipo_gasto).strip().lower()
    return MAPA_CONCEPTOS.get(tg, tg.replace("_", " ").capitalize())

# ═══════════════════════════════════════════════════════════════════
#  Paleta de colores — fiel a la foto de referencia
# ═══════════════════════════════════════════════════════════════════
C_NAVY      = "0A3A60"    # Azul corporativo oscuro (cabeceras tabla)
C_NAVY_TEXT = "FFFFFF"    # Texto sobre azul
C_HEADER_BG = "D9E1F2"   # Azul muy claro — fondo bloques info empleado
C_ROW_ALT   = "E9F0FB"   # Azul claro alternado filas de datos
C_GOLD_BG   = "FFC000"   # Dorado fuerte — subtotal
C_GOLD_LIGHT= "FFE699"   # Dorado claro — anticipo
C_LABEL_BG  = "F2F2F2"   # Gris muy claro para etiquetas de celda
C_BORDER    = "8EAADB"   # Borde azulado
C_DARK_TEXT = "0F172A"
C_GREEN     = "166534"
C_SALDO_BG  = "D6E4BC"   # Verde claro para saldo

def _thin(color=C_BORDER):
    return Side(style="thin", color=color)

def _border(color=C_BORDER):
    s = _thin(color)
    return Border(left=s, right=s, top=s, bottom=s)

def _border_all_dark():
    s = _thin("000000")
    return Border(left=s, right=s, top=s, bottom=s)

def _fill(hex_color):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

def _font(bold=False, color=C_DARK_TEXT, size=10, italic=False):
    return Font(name="Calibri", size=size, bold=bold, color=color, italic=italic)

def _al(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _set_row_height(ws, row, height):
    ws.row_dimensions[row].height = height

def _label_cell(ws, row, col, text):
    """Celda de etiqueta (fondo gris, negrita)."""
    c = ws.cell(row=row, column=col, value=text)
    c.font = _font(bold=True, size=9)
    c.fill = _fill(C_LABEL_BG)
    c.border = _border("BFBFBF")
    c.alignment = _al("left", "center")

def _value_cell(ws, row, col, value):
    """Celda de valor."""
    c = ws.cell(row=row, column=col, value=value)
    c.font = _font(size=10)
    c.border = _border("BFBFBF")
    c.alignment = _al("left", "center")
    return c

def _header_brand(ws, row_title, row_code, company_cols, meta_col_start, total_cols):
    """
    Fila de marca: compañía centrada, metadatos a la derecha.
    row_title  : fila donde va el nombre de la empresa
    row_code   : primera fila de los metadatos (Código, Versión, etc.)
    company_cols: (inicio, fin) rango de merge para nombre empresa
    meta_col_start: columna inicio metadatos
    total_cols  : total de columnas de la hoja
    """
    c1, c2 = company_cols

    # ── Fila 1: Nombre empresa ──────────────────────────────────────
    ws.merge_cells(start_row=row_title, start_column=2, end_row=row_title, end_column=c2 - 1)
    cell_company = ws.cell(row=row_title, column=2, value="GLOBAL SECURITY BANK SAS")
    cell_company.font = _font(bold=True, size=13, color=C_NAVY)
    cell_company.alignment = _al("center", "center")
    cell_company.fill = _fill("FFFFFF")

    # ── Metadatos derecha ──────────────────────────────────────────
    meta = [
        ("Código",    "OP-FR-02"),
        ("Versión",   "3"),
        ("Fecha Act", date.today().strftime("%b-%y")),
        ("Página",    "1 de 1"),
    ]
    for i, (lbl, val) in enumerate(meta):
        r = row_code + i
        ws.cell(row=r, column=meta_col_start, value=lbl).font = _font(bold=True, size=9, color=C_DARK_TEXT)
        ws.cell(row=r, column=meta_col_start).fill = _fill(C_LABEL_BG)
        ws.cell(row=r, column=meta_col_start).border = _border("BFBFBF")
        ws.cell(row=r, column=meta_col_start).alignment = _al("center")
        v_cell = ws.cell(row=r, column=meta_col_start + 1, value=val)
        v_cell.font = _font(size=9)
        v_cell.border = _border("BFBFBF")
        v_cell.alignment = _al("center")

def _table_header_row(ws, row, headers, col_start=1):
    for ci, text in enumerate(headers, start=col_start):
        c = ws.cell(row=row, column=ci, value=text)
        c.font = _font(bold=True, color=C_NAVY_TEXT, size=10)
        c.fill = _fill(C_NAVY)
        c.alignment = _al("center", "center", wrap=True)
        c.border = _border_all_dark()
        _set_row_height(ws, row, 30)

def _data_row(ws, row, values, alt=False, col_start=1):
    bg = C_ROW_ALT if alt else "FFFFFF"
    for ci, val in enumerate(values, start=col_start):
        c = ws.cell(row=row, column=ci, value=val)
        c.font = _font(size=9)
        c.fill = _fill(bg)
        c.border = _border()
        c.alignment = _al("left", "center")
    return row

def _currency_cell(ws, row, col, value, bg=None, bold=False, color=C_DARK_TEXT):
    c = ws.cell(row=row, column=col, value=float(value))
    c.font = _font(bold=bold, color=color)
    c.alignment = _al("right", "center")
    c.number_format = '_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)'
    if bg:
        c.fill = _fill(bg)
    c.border = _border_all_dark()
    return c

def _autofit(ws, extra=4, minimum=10, maximum=40):
    for col in ws.columns:
        max_len = 0
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        letter = get_column_letter(col[0].column)
        ws.column_dimensions[letter].width = min(max(max_len + extra, minimum), maximum)


# ═══════════════════════════════════════════════════════════════════
#  HELPERS DE IMAGEN Y FORMATO
# ═══════════════════════════════════════════════════════════════════
def _get_logo_image(anchor: str = "B1", width: int = 105, height: int = 90) -> Optional[object]:
    """Carga e instancia el logotipo corporativo GSB para insertarlo en openpyxl."""
    try:
        from openpyxl.drawing.image import Image as XLImage
        logo_png = os.path.join(os.path.dirname(__file__), "templates", "logo-gsb.png")
        if os.path.exists(logo_png):
            img = XLImage(logo_png)
            img.width = width
            img.height = height
            img.anchor = anchor
            return img

        if os.path.exists(TEMPLATE_PATH):
            with zipfile.ZipFile(TEMPLATE_PATH, "r") as zf:
                if "xl/media/image1.jpeg" in zf.namelist():
                    img_bytes = zf.read("xl/media/image1.jpeg")
                    img = XLImage(io.BytesIO(img_bytes))
                    img.width = width
                    img.height = height
                    img.anchor = anchor
                    return img
    except Exception:
        pass
    return None


# ═══════════════════════════════════════════════════════════════════
#  EXCEL INDEPENDIENTES
# ═══════════════════════════════════════════════════════════════════
def generar_excel_viaticos_independientes(
    usuario: Usuario,
    viaticos: List[Viatico],
    fecha_inicio=None,
    fecha_fin=None,
) -> io.BytesIO:
    """
    Genera un archivo Excel (.xlsx) con los viáticos independientes
    siguiendo el formato corporativo GSB (basado en OP-FR-02).
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Viáticos Independientes"

    TOTAL_COLS = 10   # A..J
    META_COL   = 9    # cols I, J

    # ── Columnas de ancho fijo ──────────────────────────────────
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 24
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 12
    ws.column_dimensions["I"].width = 14
    ws.column_dimensions["J"].width = 16

    # ══ FILA 1: Logo  |  Empresa  |  Metadatos ══════════════════
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 18
    ws.row_dimensions[4].height = 18

    # Logo corporativo en celda A1
    logo_img = _get_logo_image(anchor="A1", width=75, height=75)
    if logo_img:
        ws.add_image(logo_img)
    else:
        ws.merge_cells("A1:A4")
        logo_fallback = ws["A1"]
        logo_fallback.value = "🛡 GSB"
        logo_fallback.font = _font(bold=True, size=14, color=C_NAVY)
        logo_fallback.alignment = _al("center", "center")
        logo_fallback.fill = _fill("E9F0FB")
        logo_fallback.border = _border(C_NAVY)

    # Nombre empresa (cols B-H, fila 1 a 4)
    ws.merge_cells("B1:H4")
    empresa_cell = ws["B1"]
    empresa_cell.value = "GLOBAL SECURITY BANK SAS"
    empresa_cell.font = _font(bold=True, size=14, color=C_NAVY)
    empresa_cell.alignment = _al("center", "center")
    empresa_cell.fill = _fill("FFFFFF")
    empresa_cell.border = _border(C_NAVY)

    # Metadatos derecha I-J
    meta_rows = [
        ("Código",    "OP-FR-02"),
        ("Versión",   "3"),
        ("Fecha Act", date.today().strftime("%b-%y")),
        ("Página",    "1 de 1"),
    ]
    for i, (lbl, val) in enumerate(meta_rows):
        r = i + 1
        cl = ws.cell(row=r, column=9, value=lbl)
        cl.font = _font(bold=True, size=9)
        cl.fill = _fill(C_LABEL_BG)
        cl.border = _border("BFBFBF")
        cl.alignment = _al("center", "center")

        cv = ws.cell(row=r, column=10, value=val)
        cv.font = _font(size=9)
        cv.border = _border("BFBFBF")
        cv.alignment = _al("center", "center")

    # ══ BLOQUE INFO EMPLEADO (filas 6-8) ═════════════════════════
    # Fila 5: separador
    ws.row_dimensions[5].height = 6

    # Fila 6
    ws.row_dimensions[6].height = 22
    _label_cell(ws, 6, 1, "Nombres :")
    ws.merge_cells("B6:D6")
    _value_cell(ws, 6, 2, usuario.nombre or "—")
    _label_cell(ws, 6, 5, "Apellidos:")
    ws.merge_cells("F6:J6")
    _value_cell(ws, 6, 6, "—")

    # Fila 7
    ws.row_dimensions[7].height = 22
    _label_cell(ws, 7, 1, "Cédula :")
    ws.merge_cells("B7:C7")
    _value_cell(ws, 7, 2, usuario.codigo_empleado or "—")
    _label_cell(ws, 7, 4, "Período :")
    str_ini = fecha_inicio.strftime("%d/%m/%Y") if fecha_inicio else "Inicio"
    str_fin = fecha_fin.strftime("%d/%m/%Y") if fecha_fin else "Hoy"
    ws.merge_cells("E7:G7")
    _value_cell(ws, 7, 5, f"{str_ini}  →  {str_fin}")
    _label_cell(ws, 7, 8, "Fecha planilla :")
    ws.merge_cells("I7:J7")
    _value_cell(ws, 7, 9, date.today().strftime("%d/%m/%Y"))

    # Fila 8: Total ítems
    ws.row_dimensions[8].height = 22
    _label_cell(ws, 8, 1, "Total ítems :")
    _value_cell(ws, 8, 2, len(viaticos))
    _label_cell(ws, 8, 4, "Correo :")
    ws.merge_cells("E8:G8")
    _value_cell(ws, 8, 5, usuario.correo or "—")

    # ══ TABLA DE DATOS (fila 10 en adelante) ═════════════════════
    ws.row_dimensions[9].height = 6   # separador

    HEADERS = [
        "No. ítem",
        "Fecha de Viaje",
        "NIT / Identificación",
        "Razón Social",
        "Concepto",
        "Origen",
        "Destino",
        "Tiene soporte\n(si o no)",
        "Estado",
        "Valor",
    ]
    HEADER_ROW = 10
    _table_header_row(ws, HEADER_ROW, HEADERS, col_start=1)

    row_idx = HEADER_ROW + 1
    total_valor = Decimal("0.00")

    for i, v in enumerate(viaticos, start=1):
        try:
            meta = json.loads(v.descripcion or "{}")
        except Exception:
            meta = {}

        origen  = meta.get("origen", "—") or "—"
        destino = meta.get("destino", v.ciudad or "—") or "—"
        tiene_soporte = "SI" if (getattr(v, "evidencias", None) and len(v.evidencias) > 0) else "no"
        val = Decimal(str(v.valor or 0))
        estado_raw = (v.estado or "pendiente").strip().lower()
        estado_cap = estado_raw.capitalize()

        # Solo se suma al total si NO está rechazado
        if estado_raw != "rechazado":
            total_valor += val

        alt = (i % 2 == 0)
        bg = C_ROW_ALT if alt else "FFFFFF"
        row_data = [
            i,
            v.fecha.strftime("%d-%b") if isinstance(v.fecha, date) else str(v.fecha),
            v.nit_identificacion or "—",
            v.cliente or "—",
            _formatear_concepto(v.tipo_gasto),
            origen,
            destino,
            tiene_soporte,
            estado_cap,
        ]

        ws.row_dimensions[row_idx].height = 20
        for ci, cell_val in enumerate(row_data, start=1):
            c = ws.cell(row=row_idx, column=ci, value=cell_val)
            c.font = _font(size=9)
            c.fill = _fill(bg)
            c.border = _border()
            c.alignment = _al("center" if ci in (1, 2, 8, 9) else "left", "center")

            # Estilo condicional para columna Estado (col 9)
            if ci == 9:
                if estado_raw == "aprobado":
                    c.font = _font(bold=True, color="166534", size=9)
                    c.fill = _fill("E6F4EA")
                elif estado_raw == "rechazado":
                    c.font = _font(bold=True, color="C00000", size=9)
                    c.fill = _fill("FCE8E6")
                else:
                    c.font = _font(bold=True, color="B45309", size=9)
                    c.fill = _fill("FEF3C7")

        # Columna 10 (J): valor
        c_val = _currency_cell(ws, row_idx, 10, val, bg=bg)
        if estado_raw == "rechazado":
            c_val.font = _font(color="888888", size=9)

        row_idx += 1

    # ══ FILAS DE TOTALES ═════════════════════════════════════════
    ws.row_dimensions[row_idx].height = 22
    # Subtotal label (cols H-I merged)
    ws.merge_cells(start_row=row_idx, start_column=8, end_row=row_idx, end_column=9)
    sub_lbl = ws.cell(row=row_idx, column=8, value="Subtotal")
    sub_lbl.font = _font(bold=True, size=10, color="000000")
    sub_lbl.fill = _fill(C_GOLD_BG)
    sub_lbl.alignment = _al("right", "center")
    sub_lbl.border = _border_all_dark()

    # Subtotal con fórmula SUMIF que excluye rechazados
    last_data_row = row_idx - 1
    c_sub = ws.cell(
        row=row_idx,
        column=10,
        value=f'=SUMIF(I11:I{last_data_row}, "<>Rechazado", J11:J{last_data_row})' if last_data_row >= 11 else total_valor,
    )
    c_sub.font = _font(bold=True, color=C_DARK_TEXT, size=10)
    c_sub.alignment = _al("right", "center")
    c_sub.number_format = '_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)'
    c_sub.fill = _fill(C_GOLD_BG)
    c_sub.border = _border_all_dark()

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream


# ═══════════════════════════════════════════════════════════════════
#  EXCEL ASIGNACIÓN
# ═══════════════════════════════════════════════════════════════════
def generar_excel_viaticos_asignacion(
    asignacion: Asignacion,
    viaticos: List[Viatico],
) -> io.BytesIO:
    """
    Genera un archivo Excel (.xlsx) con los viáticos asociados a una Asignación
    utilizando como plantilla base el archivo oficial FORMATO LEGALIZACION VIATICOS.xlsx.
    """
    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    ws = wb.active

    # ── Inserción del Logotipo Corporativo GSB ────────────────────────────────
    ws._images = []  # limpiar referencias heredadas para evitar duplicados o fallos
    logo_img = _get_logo_image(anchor="B1", width=105, height=90)
    if logo_img:
        ws.add_image(logo_img)
    # ─────────────────────────────────────────────────────────────────────────

    tecnico = asignacion.tecnico
    nombre_tecnico = tecnico.nombre if tecnico else "—"
    cedula_tecnico = tecnico.codigo_empleado if tecnico else "—"
    nombres, apellidos = dividir_nombres_apellidos(nombre_tecnico)
    anticipo_num = float(asignacion.monto_anticipo or 0)

    # ── Encabezados / Bloque Info ────────────────────────────────
    ws["D6"] = nombres
    ws["G6"] = apellidos
    ws["C7"] = cedula_tecnico

    aprobados_count = len([v for v in viaticos if str(v.estado).lower() == "aprobado"])
    # Celda F7: mostrar las órdenes de servicio (OT) reales de los viáticos
    ots_unicas = list(dict.fromkeys(
        v.ot.strip() for v in viaticos if v.ot and v.ot.strip()
    ))
    ws["F7"] = ", ".join(ots_unicas) if ots_unicas else str(aprobados_count)
    ws["J7"] = date.today().strftime("%d/%m/%y")

    if asignacion.fecha_inicio:
        ws["C8"] = _formato_fecha_anticipo(asignacion.fecha_inicio)
    else:
        ws["C8"] = "—"

    ws["G8"] = "SI____ NO____"
    ws["I8"] = "SI____ NO____"
    ws["K8"] = anticipo_num

    # ── Encabezado Columna Estado (L10) ──────────────────────────
    c_l10 = ws.cell(row=10, column=12, value="Estado")
    c_k10 = ws.cell(row=10, column=11)
    if c_k10.has_style:
        c_l10.font = copy(c_k10.font)
        c_l10.fill = copy(c_k10.fill)
        c_l10.border = copy(c_k10.border)
        c_l10.alignment = copy(c_k10.alignment)
    else:
        c_l10.font = _font(bold=True, size=10, color=C_NAVY_TEXT)
        c_l10.fill = _fill(C_NAVY)
        c_l10.border = _border_all_dark()
        c_l10.alignment = _al("center", "center")
    ws.column_dimensions["L"].width = 15

    # ── Tabla de ítems ──────────────────────────────────────────
    total_items = len(viaticos)
    base_slots = 14  # filas 11 a 24 en la plantilla original

    if total_items > base_slots:
        extra = total_items - base_slots
        ws.insert_rows(25, amount=extra)
        for r in range(25, 25 + extra):
            for c in range(1, 13):
                src = ws.cell(row=24, column=c)
                dst = ws.cell(row=r, column=c)
                if src.has_style:
                    dst.font = copy(src.font)
                    dst.border = copy(src.border)
                    dst.fill = copy(src.fill)
                    dst.number_format = copy(src.number_format)
                    dst.alignment = copy(src.alignment)
        last_item_row = 24 + extra
    else:
        last_item_row = 24

    for idx, v in enumerate(viaticos):
        r = 11 + idx
        try:
            meta = json.loads(v.descripcion or "{}")
        except Exception:
            meta = {}

        nit = v.nit_identificacion or meta.get("nit") or "—"
        razon = meta.get("razon_social") or v.cliente or asignacion.cliente or "—"
        concepto = _formatear_concepto(v.tipo_gasto)
        oficina = asignacion.empresa or asignacion.ciudad or "—"
        origen = meta.get("origen") or "—"
        destino = meta.get("destino") or v.ciudad or asignacion.ciudad or "—"
        tiene_soporte = (
            "SI"
            if (
                meta.get("tiene_soporte") is True
                or (getattr(v, "evidencias", None) and len(v.evidencias) > 0)
            )
            else "NO"
        )
        val = float(v.valor or 0)
        estado_raw = (v.estado or "pendiente").strip().lower()
        estado_cap = estado_raw.capitalize()

        ws.cell(row=r, column=2, value=idx + 1)
        ws.cell(row=r, column=3, value=_formato_fecha_viaje(v.fecha))
        ws.cell(row=r, column=4, value=nit)
        ws.cell(row=r, column=5, value=razon)
        ws.cell(row=r, column=6, value=concepto)
        ws.cell(row=r, column=7, value=oficina)
        ws.cell(row=r, column=8, value=origen)
        ws.cell(row=r, column=9, value=destino)
        ws.cell(row=r, column=10, value=tiene_soporte)
        
        c_val = ws.cell(row=r, column=11, value=val)
        c_val.number_format = '_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)'

        # Columna 12 (L): Estado de aprobación/rechazo
        c_est = ws.cell(row=r, column=12, value=estado_cap)
        if ws.cell(row=r, column=11).has_style:
            c_est.border = copy(ws.cell(row=r, column=11).border)
        else:
            c_est.border = _border()
        c_est.alignment = _al("center", "center")

        if estado_raw == "aprobado":
            c_est.font = _font(bold=True, size=9, color="166534")
            c_est.fill = _fill("E6F4EA")
        elif estado_raw == "rechazado":
            c_est.font = _font(bold=True, size=9, color="C00000")
            c_est.fill = _fill("FCE8E6")
            # Texto atenuado para el valor rechazado
            c_val.font = _font(size=9, color="888888")
        else:
            c_est.font = _font(bold=True, size=9, color="B45309")
            c_est.fill = _fill("FEF3C7")

    # ── Totales y Fórmulas ──────────────────────────────────────
    subtotal_row = last_item_row + 1
    ant_row = subtotal_row + 1
    gsb_row = ant_row + 1
    tec_row = gsb_row + 1
    tot_row = tec_row + 1

    ws.cell(row=subtotal_row, column=10, value="Subtotal")
    # Subtotal dinámico: SUMIF suma únicamente los que NO digan 'Rechazado'
    ws.cell(row=subtotal_row, column=11, value=f'=SUMIF(L11:L{last_item_row}, "<>Rechazado", K11:K{last_item_row})')

    ws.cell(row=ant_row, column=10, value="valor del anticipo")
    ws.cell(row=ant_row, column=11, value="=K8")

    ws.cell(row=gsb_row, column=10, value="saldo a favor GSB")
    ws.cell(row=gsb_row, column=11, value=f"=IF(K{ant_row}>K{subtotal_row}, K{ant_row}-K{subtotal_row}, 0)")

    ws.cell(row=tec_row, column=10, value="saldo a favor TECNICO")
    ws.cell(row=tec_row, column=11, value=f"=IF(K{subtotal_row}>K{ant_row}, K{subtotal_row}-K{ant_row}, 0)")

    ws.cell(row=tot_row, column=10, value="TOTAL LEGALIZADO")
    ws.cell(row=tot_row, column=11, value=f"=K{subtotal_row}")

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream


# ═══════════════════════════════════════════════════════════════════
#  EXCEL TALENTO HUMANO (PLANILLA CORPORATIVA)
# ═══════════════════════════════════════════════════════════════════
def generar_excel_talento_humano(empleados_data: list) -> io.BytesIO:
    """
    Genera un archivo Excel (.xlsx) con el consolidado del personal de
    Talento Humano siguiendo el diseño corporativo de Global Security Bank.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Talento Humano"

    # ── Columnas de ancho fijo ──────────────────────────────────
    column_widths = {
        "A": 8,   # No.
        "B": 28,  # Nombre Completo
        "C": 15,  # Cédula
        "D": 16,  # Código Empleado
        "E": 22,  # Cargo
        "F": 18,  # Área
        "G": 28,  # Correo Electrónico
        "H": 16,  # Teléfono
        "I": 16,  # Ciudad
        "J": 14,  # Fecha Ingreso
        "K": 20,  # Tipo de Contrato
        "L": 22,  # Jefe Inmediato
        "M": 14,  # Estado Laboral
        "N": 18,  # Salario
        "O": 16,  # Docs Cargados
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # ══ FILA 1-4: Logo | Empresa | Metadatos ════════════════════
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 18
    ws.row_dimensions[4].height = 18

    # Logo corporativo
    logo_img = _get_logo_image(anchor="A1", width=75, height=75)
    if logo_img:
        ws.add_image(logo_img)
    else:
        ws.merge_cells("A1:A4")
        logo_cell = ws["A1"]
        logo_cell.value = "🛡 GSB"
        logo_cell.font = _font(bold=True, size=14, color=C_NAVY)
        logo_cell.alignment = _al("center", "center")
        logo_cell.fill = _fill("E9F0FB")
        logo_cell.border = _border(C_NAVY)

    # Nombre empresa
    ws.merge_cells("B1:M4")
    empresa_cell = ws["B1"]
    empresa_cell.value = "GLOBAL SECURITY BANK SAS — GESTIÓN DE TALENTO HUMANO"
    empresa_cell.font = _font(bold=True, size=13, color=C_NAVY)
    empresa_cell.alignment = _al("center", "center")
    empresa_cell.fill = _fill("FFFFFF")
    empresa_cell.border = _border(C_NAVY)

    # Metadatos N-O
    meta_rows = [
        ("Código", "TH-FR-01"),
        ("Versión", "1"),
        ("Fecha Gen", date.today().strftime("%d/%m/%Y")),
        ("Total Empleados", str(len(empleados_data))),
    ]
    for i, (lbl, val) in enumerate(meta_rows):
        r = i + 1
        cl = ws.cell(row=r, column=14, value=lbl)
        cl.font = _font(bold=True, size=9)
        cl.fill = _fill(C_LABEL_BG)
        cl.border = _border("BFBFBF")
        cl.alignment = _al("center", "center")

        cv = ws.cell(row=r, column=15, value=val)
        cv.font = _font(size=9)
        cv.border = _border("BFBFBF")
        cv.alignment = _al("center", "center")

    # Separador
    ws.row_dimensions[5].height = 8

    # ══ TABLA DE ENCABEZADOS ═════════════════════════════════════
    HEADERS = [
        "No.",
        "Nombre completo",
        "Cédula",
        "Código empleado",
        "Cargo",
        "Área",
        "Correo electrónico",
        "Teléfono",
        "Ciudad",
        "Fecha ingreso",
        "Tipo contrato",
        "Jefe inmediato",
        "Estado laboral",
        "Salario (COP)",
        "Documentación",
    ]
    HEADER_ROW = 6
    _table_header_row(ws, HEADER_ROW, HEADERS, col_start=1)

    row_idx = HEADER_ROW + 1
    total_salarios = Decimal("0.00")

    for i, emp in enumerate(empleados_data, start=1):
        alt = (i % 2 == 0)
        bg = C_ROW_ALT if alt else "FFFFFF"

        nombre = emp.get("nombre") or "—"
        cedula = emp.get("cedula") or emp.get("codigo_empleado") or "—"
        codigo = emp.get("codigo_empleado") or "—"
        cargo = emp.get("cargo") or "Técnico Instalador"
        area = emp.get("area") or "Instalaciones"
        correo = emp.get("correo") or "—"
        telefono = emp.get("telefono") or "—"
        ciudad = emp.get("ciudad") or "—"
        
        f_ingreso = emp.get("fecha_ingreso")
        if isinstance(f_ingreso, (date, datetime)):
            str_ingreso = f_ingreso.strftime("%d/%m/%Y")
        elif f_ingreso:
            str_ingreso = str(f_ingreso)
        else:
            str_ingreso = "—"

        tipo_contrato = emp.get("tipo_contrato") or "Término indefinido"
        jefe = emp.get("jefe_inmediato") or "Carlos Ramírez"
        estado_raw = (emp.get("estado_laboral") or "activo").strip().lower()
        estado_cap = estado_raw.replace("_", " ").capitalize()
        
        salario_val = Decimal(str(emp.get("salario") or 0))
        total_salarios += salario_val

        docs_info = f"{emp.get('documentos_cargados', 0)} / {emp.get('documentos_totales', 6)}"

        row_data = [
            i,
            nombre,
            cedula,
            codigo,
            cargo,
            area,
            correo,
            telefono,
            ciudad,
            str_ingreso,
            tipo_contrato,
            jefe,
            estado_cap,
        ]

        ws.row_dimensions[row_idx].height = 20
        for ci, cell_val in enumerate(row_data, start=1):
            c = ws.cell(row=row_idx, column=ci, value=cell_val)
            c.font = _font(size=9)
            c.fill = _fill(bg)
            c.border = _border()
            c.alignment = _al("center" if ci in (1, 3, 4, 8, 9, 10, 13) else "left", "center")

            # Estilo para estado
            if ci == 13:
                if estado_raw == "activo":
                    c.font = _font(bold=True, color="166534", size=9)
                    c.fill = _fill("E6F4EA")
                elif estado_raw == "en_capacitacion":
                    c.font = _font(bold=True, color="B45309", size=9)
                    c.fill = _fill("FEF3C7")
                else:
                    c.font = _font(bold=True, color="C00000", size=9)
                    c.fill = _fill("FCE8E6")

        # Columna 14 (N): Salario
        _currency_cell(ws, row_idx, 14, salario_val, bg=bg)

        # Columna 15 (O): Documentación
        c_doc = ws.cell(row=row_idx, column=15, value=docs_info)
        c_doc.font = _font(size=9, bold=True)
        c_doc.fill = _fill(bg)
        c_doc.border = _border()
        c_doc.alignment = _al("center", "center")

        row_idx += 1

    # Fila de Totales
    ws.row_dimensions[row_idx].height = 22
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=13)
    lbl_tot = ws.cell(row=row_idx, column=1, value="Masa Salarial Total (COP)")
    lbl_tot.font = _font(bold=True, size=10, color="000000")
    lbl_tot.fill = _fill(C_GOLD_BG)
    lbl_tot.alignment = _al("right", "center")
    lbl_tot.border = _border_all_dark()

    c_tot = ws.cell(
        row=row_idx,
        column=14,
        value=f"=SUM(N7:N{row_idx - 1})" if row_idx > 7 else float(total_salarios),
    )
    c_tot.font = _font(bold=True, color=C_DARK_TEXT, size=10)
    c_tot.alignment = _al("right", "center")
    c_tot.number_format = '_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)'
    c_tot.fill = _fill(C_GOLD_BG)
    c_tot.border = _border_all_dark()

    ws.cell(row=row_idx, column=15).fill = _fill(C_GOLD_BG)
    ws.cell(row=row_idx, column=15).border = _border_all_dark()

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream

